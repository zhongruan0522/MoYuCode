#!/usr/bin/env python3
"""
Excel Handler Tool
Read, write, and manipulate Excel files.
Based on: https://github.com/python-excel/xlrd, openpyxl

Usage:
    python excel_handler.py read --input data.xlsx --output data.json
    python excel_handler.py create --input data.json --output report.xlsx
    python excel_handler.py convert --input data.xlsx --output data.csv
    python excel_handler.py merge --inputs file1.xlsx,file2.xlsx --output merged.xlsx

Requirements:
    pip install openpyxl pandas
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("Error: Required packages missing. Install with: pip install openpyxl pandas", file=sys.stderr)
    sys.exit(1)


def read_excel(
    input_path: str,
    output_path: str = None,
    sheet_name: str = None,
    as_dict: bool = True,
) -> dict | list:
    """
    Read Excel file and optionally export to JSON.
    
    Args:
        input_path: Path to Excel file
        output_path: Optional path to save JSON output
        sheet_name: Specific sheet to read (default: all sheets)
        as_dict: Return as dict with sheet names as keys
    
    Returns:
        Dictionary or list of data from Excel
    """
    try:
        if sheet_name:
            df = pd.read_excel(input_path, sheet_name=sheet_name)
            data = df.to_dict(orient='records')
        else:
            # Read all sheets
            xlsx = pd.ExcelFile(input_path)
            data = {}
            for sheet in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet)
                # Convert NaN to None for JSON compatibility
                df = df.where(pd.notnull(df), None)
                data[sheet] = df.to_dict(orient='records')
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2, ensure_ascii=False, default=str)
            print(f"✓ Exported to JSON: {output_path}")
        
        return data
    
    except Exception as e:
        print(f"Error reading Excel: {e}", file=sys.stderr)
        return None


def create_excel(
    input_path: str,
    output_path: str,
    sheet_name: str = "Sheet1",
    header_style: bool = True,
) -> bool:
    """
    Create Excel file from JSON or CSV data.
    
    Args:
        input_path: Path to JSON or CSV file
        output_path: Path for output Excel file
        sheet_name: Name for the worksheet
        header_style: Apply styling to header row
    
    Returns:
        True if successful
    """
    try:
        input_file = Path(input_path)
        
        # Load data based on file type
        if input_file.suffix.lower() == '.json':
            with open(input_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            # Handle dict with multiple sheets
            if isinstance(data, dict) and not all(isinstance(v, (str, int, float, bool, type(None))) for v in data.values()):
                # Multiple sheets
                wb = Workbook()
                wb.remove(wb.active)  # Remove default sheet
                
                for sheet_name, sheet_data in data.items():
                    if isinstance(sheet_data, list):
                        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit
                        df = pd.DataFrame(sheet_data)
                        _write_dataframe_to_sheet(ws, df, header_style)
                
                wb.save(output_path)
                print(f"✓ Created Excel with {len(data)} sheets: {output_path}")
                return True
            else:
                # Single sheet
                df = pd.DataFrame(data if isinstance(data, list) else [data])
        
        elif input_file.suffix.lower() == '.csv':
            df = pd.read_csv(input_path)
        
        else:
            print(f"Error: Unsupported input format: {input_file.suffix}", file=sys.stderr)
            return False
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        
        _write_dataframe_to_sheet(ws, df, header_style)
        
        wb.save(output_path)
        print(f"✓ Created Excel: {output_path}")
        return True
    
    except Exception as e:
        print(f"Error creating Excel: {e}", file=sys.stderr)
        return False


def _write_dataframe_to_sheet(ws, df: pd.DataFrame, header_style: bool = True):
    """Write DataFrame to worksheet with optional styling."""
    # Write headers
    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        if header_style:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            # Handle NaN/None
            if pd.isna(value):
                value = None
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def convert_excel(
    input_path: str,
    output_path: str,
    sheet_name: str = None,
) -> bool:
    """
    Convert Excel to CSV or other formats.
    
    Args:
        input_path: Path to Excel file
        output_path: Path for output file
        sheet_name: Sheet to convert (default: first sheet)
    
    Returns:
        True if successful
    """
    try:
        output_format = Path(output_path).suffix.lower()
        
        if sheet_name:
            df = pd.read_excel(input_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(input_path)
        
        if output_format == '.csv':
            df.to_csv(output_path, index=False, encoding='utf-8')
        elif output_format == '.json':
            df.to_json(output_path, orient='records', indent=2, force_ascii=False)
        elif output_format == '.html':
            df.to_html(output_path, index=False)
        elif output_format == '.md':
            df.to_markdown(output_path, index=False)
        else:
            print(f"Error: Unsupported output format: {output_format}", file=sys.stderr)
            return False
        
        print(f"✓ Converted: {input_path} -> {output_path}")
        return True
    
    except Exception as e:
        print(f"Error converting Excel: {e}", file=sys.stderr)
        return False


def merge_excel(
    input_paths: list[str],
    output_path: str,
    combine_sheets: bool = False,
) -> bool:
    """
    Merge multiple Excel files.
    
    Args:
        input_paths: List of Excel file paths
        output_path: Path for merged output
        combine_sheets: If True, combine all into one sheet
    
    Returns:
        True if successful
    """
    try:
        wb = Workbook()
        wb.remove(wb.active)
        
        all_data = []
        
        for input_path in input_paths:
            xlsx = pd.ExcelFile(input_path)
            file_name = Path(input_path).stem
            
            for sheet in xlsx.sheet_names:
                df = pd.read_excel(xlsx, sheet_name=sheet)
                
                if combine_sheets:
                    df['_source_file'] = file_name
                    df['_source_sheet'] = sheet
                    all_data.append(df)
                else:
                    # Create unique sheet name
                    sheet_title = f"{file_name}_{sheet}"[:31]
                    ws = wb.create_sheet(title=sheet_title)
                    _write_dataframe_to_sheet(ws, df, header_style=True)
        
        if combine_sheets and all_data:
            combined_df = pd.concat(all_data, ignore_index=True)
            ws = wb.create_sheet(title="Combined")
            _write_dataframe_to_sheet(ws, combined_df, header_style=True)
        
        wb.save(output_path)
        print(f"✓ Merged {len(input_paths)} files: {output_path}")
        return True
    
    except Exception as e:
        print(f"Error merging Excel files: {e}", file=sys.stderr)
        return False


def main():
    parser = argparse.ArgumentParser(description="Excel file handler")
    subparsers = parser.add_subparsers(dest="command", help="Commands")
    
    # Read command
    read_parser = subparsers.add_parser("read", help="Read Excel to JSON")
    read_parser.add_argument("--input", "-i", required=True, help="Input Excel file")
    read_parser.add_argument("--output", "-o", help="Output JSON file")
    read_parser.add_argument("--sheet", "-s", help="Specific sheet name")
    
    # Create command
    create_parser = subparsers.add_parser("create", help="Create Excel from JSON/CSV")
    create_parser.add_argument("--input", "-i", required=True, help="Input JSON/CSV file")
    create_parser.add_argument("--output", "-o", required=True, help="Output Excel file")
    create_parser.add_argument("--sheet", "-s", default="Sheet1", help="Sheet name")
    create_parser.add_argument("--no-style", action="store_true", help="Skip header styling")
    
    # Convert command
    convert_parser = subparsers.add_parser("convert", help="Convert Excel to CSV/JSON")
    convert_parser.add_argument("--input", "-i", required=True, help="Input Excel file")
    convert_parser.add_argument("--output", "-o", required=True, help="Output file (csv/json/html/md)")
    convert_parser.add_argument("--sheet", "-s", help="Specific sheet name")
    
    # Merge command
    merge_parser = subparsers.add_parser("merge", help="Merge multiple Excel files")
    merge_parser.add_argument("--inputs", "-i", required=True, help="Comma-separated input files")
    merge_parser.add_argument("--output", "-o", required=True, help="Output Excel file")
    merge_parser.add_argument("--combine", "-c", action="store_true", help="Combine all into one sheet")
    
    args = parser.parse_args()
    
    if not args.command:
        parser.print_help()
        sys.exit(1)
    
    success = False
    
    if args.command == "read":
        result = read_excel(args.input, args.output, args.sheet)
        success = result is not None
        if success and not args.output:
            print(json.dumps(result, indent=2, ensure_ascii=False, default=str))
    
    elif args.command == "create":
        success = create_excel(args.input, args.output, args.sheet, not args.no_style)
    
    elif args.command == "convert":
        success = convert_excel(args.input, args.output, args.sheet)
    
    elif args.command == "merge":
        input_files = [f.strip() for f in args.inputs.split(",")]
        success = merge_excel(input_files, args.output, args.combine)
    
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
